from django.shortcuts import render, get_object_or_404, redirect
from campaigns.utils import PaymentStatus
from events.models import TicketOrderModel, EventModel, TicketModel
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from django.http import HttpResponse
from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import get_template
import logging

logger = logging.getLogger("events")

@login_required
def manage_ticket_orders(request):
    orders = TicketOrderModel.objects.filter(buyer = request.user)
    return render(request, "events/orders/manage/orders.html", {"orders": orders})

@login_required
def manage_ticket_order(request, order_id):
    order_queryset = TicketOrderModel.objects.filter(buyer = request.user).select_related("event").prefetch_related("tickets")
    order = get_object_or_404(order_queryset, id=order_id)
    return render(request, "events/orders/manage/order.html", {"order": order})

@login_required
def manage_events(request):
    events = EventModel.objects.filter(organiser = request.user)
    return render(request, "events/event/manage/events.html", {"events": events})

@login_required
def manage_event(request, event_slug):
    events = EventModel.objects.filter(organiser = request.user).prefetch_related("tickettypes", "ticket_orders")
    event = get_object_or_404(events, slug=event_slug)
    total_sales = sum([order.total_price for order in event.ticket_orders.filter(paid = PaymentStatus.PAID)])
    return render(request, "events/event/manage/event.html", {"event": event, "sale": total_sales})

@login_required
def generate_ticket(request, order_id, ticket_id):
    order = get_object_or_404(TicketOrderModel, buyer=request.user, id=order_id, paid=PaymentStatus.PAID)
    ticket = get_object_or_404(TicketModel, ticket_order=order, id=ticket_id)

    domain = get_current_site(request).domain
    protocol = "https" if request.is_secure() else "http"

    try:
        template = get_template("ticket/ticket_new.html")
        context = {"ticket": ticket, 
                    "event": order.event, 
                    "buyer_full_name": order.buyer.get_full_name(),
                    "order_number": order.order_number,
                    "created": order.created,
                    "domain": domain, 
                    "protocol": protocol}
        
        render_template = template.render(context)
        pdf_file = HTML(string=render_template).write_pdf()
            
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{order.order_number}_ticket.pdf"'
        
        return response
    
    except Exception as ex:
        logger.error(ex)
        messages.error(request, "Sorry there was an error trying to generate your ticket")
        return redirect("events:manage-ticket-order", order_id=order.id)

@login_required
def generate_guest_list(request, event_id):
    event = get_object_or_404(EventModel, organiser=request.user, id=event_id)
    domain = get_current_site(request).domain

    protocol = "https" if request.is_secure() else "http"
    orders = TicketOrderModel.objects.filter(event= event, paid=PaymentStatus.PAID).prefetch_related("tickets")
    if orders.count() < 1:
        messages.warning(request, "Sorry You cannot generate guest list for now")
        return redirect("events:manage-event", event_slug=event.slug)
    
    try:
        template = get_template("ticket/guests.html")
        context = {"orders": orders, 
                    "event": event,
                    "domain": domain, 
                    "protocol": protocol}
        
        render_template = template.render(context)
        pdf_file = HTML(string=render_template).write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{event.title}_Guest_List.pdf"'

        return response
    
    except Exception as ex:
        logger.error(ex)
        messages.error(request, "Sorry there was an error trying to generate your guest list")
        return redirect("events:manage-event", event_slug=event.slug)

@login_required
def generate_rsvp_excel(request, event_id):
    
    event = get_object_or_404(EventModel, organiser=request.user, id=event_id)
    orders = TicketOrderModel.objects.filter(event= event).prefetch_related("tickets")
    if orders.count() < 5:
        messages.error(request, "Sorry You cannot generate guest list for now")
        return redirect("events:manage-event", event_slug=event.slug)
    workbook = Workbook()
    sheet = workbook.active

    title_cell = sheet.cell(row=1, column=1, value=event.title)
    title_cell.font = Font(bold=True, color="2b3245")

    headers = ['Ticket No', 'Ticket type', 'Full Name', 'Email Address', '# of Guests', 'Accept']
    sheet.append(headers)

    
    header_fill = PatternFill(start_color="f2f3fb", end_color="f2f3fb", fill_type="solid")
    header_font = Font(bold=True, color="281851")

    for cell in sheet['1:1']:
        cell.fill = header_fill
        cell.font = header_font
    
    for order in orders:
        for ticket in order.tickets.all():
            row_data = [
                    ticket.barcode_value,
                    ticket.guest_full_name,
                    ticket.ticket_type.title,
                    ticket.guest_email,
                    ticket.quantity,
                    ''  # Leave the Accept column empty for user input
                ]
            sheet.append(row_data)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{order.event.title}_RSVP.xlsx"'

    workbook.save(response)

    return response